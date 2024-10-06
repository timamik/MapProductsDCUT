from openpyxl import load_workbook
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Загружаем Excel файл
wb = load_workbook('БазаШтВыручкаАдресДоставки.xlsx')
ws = wb.active

# Инициализируем структуру
structure = {}
current_structure_name = None
current_address = None

# Обрабатываем строки файла
for row in ws.iter_rows(min_row=1, max_col=4):
    cell = row[0]
    first_col_font_size = cell.font.size
    first_col_font_color = cell.font.color.rgb if cell.font.color else None
    third_col_value = row[2].value  # Количество штук
    fourth_col_value = row[3].value  # Выручка

    # 1. Если шрифт 10 в первой колонке — это название структуры
    if first_col_font_size == 10:
        current_structure_name = cell.value
        structure[current_structure_name] = {}

    # 2. Если шрифт 8 и цвет черный в первой колонке — это адрес
    elif first_col_font_size == 8 and first_col_font_color == 'FF000000':  # Черный цвет
        current_address = cell.value
        if current_address not in structure[current_structure_name]:
            # Инициализируем массив для количества и выручки
            structure[current_structure_name][current_address] = [0, 0]

        # 3. Суммируем количество (колонка 3)
        structure[current_structure_name][current_address][0] += third_col_value or 0
        
        # 4. Суммируем выручку (колонка 4)
        structure[current_structure_name][current_address][1] += fourth_col_value or 0

# Записываем структуру в текстовый файл
with open('output_structure.txt', 'w', encoding='utf-8') as file:
    for structure_name, addresses in structure.items():
        # Пишем название структуры
        file.write(f"{structure_name};\n")
        
        # Пишем адреса и суммы для каждого адреса
        for address, values in addresses.items():
            quantity = values[0]
            revenue = values[1]
            file.write(f"{address}; {quantity}; {revenue}\n")
        
        file.write("\n")

print("Структура успешно сохранена в 'output_structure.txt'")
